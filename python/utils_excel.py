from openpyxl.styles import PatternFill, Font
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
import ipywidgets as widgets
import base64
from python.fuzzy import unique_list
import pandas as pd

def export_pm_amounts_to_excel(pm_names, grouped_by_pm_name, bank_key, amount_key,
                               output_path, pm_name_to_parking, pm_name_to_pm_sarl_names, pm_sarl_to_sum):
    max_bank_name = 0
    for name in pm_names:
        entries = grouped_by_pm_name.get(name, [])
        all_bank_names = unique_list([x[bank_key] for x in entries])
        if len(all_bank_names) > max_bank_name:
            max_bank_name = len(all_bank_names)

    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Amounts"
    cols = ["Name from A.xlsx", "Amount from C.xlsx", "Expected amount from A.xlsx", "Matched Name from B.xlsx"]
    for i in range(max_bank_name):
        cols.append(f"Matched name from C.xlsx {i + 1}")
    ws.append(cols)

    pm_name_to_pm_sarl_names_joined = {k: "|".join(v) for k, v in pm_name_to_pm_sarl_names.items()}

    row_red_flags = []

    for row_idx, name in enumerate(pm_names, start=2):
        parking_expected = pm_name_to_parking.get(name, 0)

        if name in pm_sarl_to_sum:
            parking_expected += pm_sarl_to_sum[name]
        else:
            for sarl_name in pm_name_to_pm_sarl_names.get(name, []):
                parking_expected += pm_sarl_to_sum.get(sarl_name, 0)

        entries = grouped_by_pm_name.get(name, [])
        all_bank_names = unique_list([trim(x[bank_key]) for x in entries])
        if not all_bank_names:
            all_bank_names = ["-- NO MATCH"]

        all_bank_names += [" "] * (max_bank_name - len(all_bank_names))

        total_amount = sum(to_float(x.get(amount_key, 0)) for x in entries)
        pm_name_sarl = pm_name_to_pm_sarl_names_joined.get(name)

        row = [name, total_amount, parking_expected, trim(pm_name_sarl)] + all_bank_names
        ws.append(row)

        is_red = mismatch(total_amount, parking_expected)
        row_red_flags.append(is_red)

        if is_red:
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            red_font = Font(color="9C0006", bold=True)
            for col_idx in range(1, len(row) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = red_fill
                cell.font = red_font

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 4

    wb.save(output_path)

    return row_red_flags

def trim(s):
    return s.strip() if s else None

def mismatch(a, b):
    return abs(a - b) >= 1

def create_download_button(filename):
    with open(filename, 'rb') as f:
        file_bytes = f.read()

    b64 = base64.b64encode(file_bytes).decode()
    payload = f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}"

    # Reduced padding and font size by about 20%
    button_html = f'''
    <a download="paymatch.xlsx" href="{payload}" 
       style="
         display: inline-block;
         padding: 8px 16px;
         font-size: 12.8px;
         font-weight: bold;
         color: white;
         background-color: red;
         border-radius: 5px;
         text-align: center;
         text-decoration: none;
         cursor: pointer;
       ">
       Download Excel
    </a>
    '''
    return widgets.HTML(value=button_html)


def parse_date(x):
    return pd.to_datetime(x, dayfirst=True, errors="coerce")


def to_float(s):
    if isinstance(s, float):
        return s / 100.0

    return float(s.replace(',', '.'))